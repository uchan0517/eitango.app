#やりたいこと:レベル別に分けて選べるようにする.左全部できたら見た目を作る。
import openpyxl
import random
import streamlit as st
import gtts
import pandas as pd
import time
global sel_Q

    
#初期化
st.session_state.page = ""
if 'matigai' not in st.session_state:
    st.session_state.matigai = 0
if 'seikai' not in st.session_state:
    st.session_state.seikai = 0
if "my_text" not in st.session_state:
    st.session_state.my_text = ""
if "answer" not in st.session_state:
    st.session_state.answer = ""
if "check" not in st.session_state:
    st.session_state.check = False
if "count" not in st.session_state:
    st.session_state.count = ""
if 'page' not in st.session_state:
    st.session_state.page = ""

#リセット
if st.button("リセット"):
    st.session_state.count = ""
    st.session_state.matigai = 0
    st.session_state.seikai = 0
#結果の表示
if st.session_state.count == 10:
    st.session_state.count = ""
    df = pd.DataFrame({
    '正解': [st.session_state.seikai],
    '間違い':[st.session_state.matigai]})
    st.dataframe(df)
#ホーム画面とサイドバー処理の場所
with st.sidebar:
    sel_Q = st.selectbox(
        "問題を解く", 
        ("使い方","中学一年生レベル","中学二年生レベル","中学三年生レベル")
    )
if st.session_state.page == "":
    st.session_state.page = sel_Q


if st.session_state.page == '使い方':
    wc = "使い方"
    st.title("これは英単語テストのソフトです")
    st.subheader("ここでは使い方を説明します")
    video_file = open("説明動画.mp4")
    st.video("説明動画.mp4")
    st.subheader("実際にしてみましょう\n終わったら左の使い方をクリックして好きなレベルを選んでください")
    st.session_state.jp = ""
    st.session_state.en = ""
    st.session_state.ave = ""
elif st.session_state.page == '中学一年生レベル':
    wc = '中一'
    st.session_state.jp = ""
    st.session_state.en = ""
    st.session_state.ave = ""
elif st.session_state.page == '中学二年生レベル':
    wc = '中ニ'
    st.session_state.jp = ""
    st.session_state.en = ""
    st.session_state.ave = ""
elif st.session_state.page == '中学三年生レベル':
    wc = '中三'
    st.session_state.jp = ""
    st.session_state.en = ""
    st.session_state.ave = ""

#英語の問題のプログラム
def QandA(English,Japanese):
    global Q
#問題つくり
    if st.session_state.count == "":
        st.session_state.mondai = random.sample(li,10)
        st.session_state.count = 0
    #st.session_state.count += 1
    print(st.session_state.mondai)
    print(st.session_state.count)
    Q = st.session_state.mondai[st.session_state.count]
    return Japanese[Q],English[Q]
    
def C(An,English):
    global Q
#正誤判定
    a = ws[f'C{Q+1}'].value
    if a == None:
        a = 0
    b = ws[f'D{Q+1}'].value
    if b == None:
        b = 0
#確率の計算
    if An == English[Q]:
        a+=1
        ws[f'C{Q+1}'] = a
    else:
        b+=1
        ws[f'D{Q+1}'] = b

    ave = f"{int((a)/(a+b)*100)}%"
    ws[f'E{Q+1}'] = ave
    wb.save('英単語.xlsx')
    #return ave

def submit():
    st.session_state.my_text = st.session_state.widget
    st.session_state.widget = ""


#ファイルの取り込み
wb = openpyxl.load_workbook('英単語.xlsx')
ws = wb[wc]
#英語の単語のリスト化
En = ws["A"]
English = [e.value for e in En if e.value != None]
if st.session_state.en == "":
    st.session_state.en = English
#日本語の単語のリスト化
Jp = ws["B"]
Japanese = [j.value for j in Jp if j.value != None]
if st.session_state.jp == "":
    st.session_state.jp = Japanese
#正解率
Ave = ws["E"]
Average = [j.value if j.value !=None else "0%" for j in Ave]
if st.session_state.ave == "":
    st.session_state.ave = Average
wb.close()
li = list(range(0,(len(English))))


def check():
    An=st.session_state.my_text
    A=st.session_state.answer
    st.session_state.check = True
    if An.lower() == A.lower():
        kotae(1)
        st.session_state.seikai += 1
    else:
        kotae(2)
        st.session_state.matigai += 1
    #if st.session_state.check:
    C(An,English)
    An = None
    #time.sleep(1)
    st.session_state.my_text = ""
    st.session_state.answer = ""
    st.session_state.check = False



@st.dialog("答え合わせ")#こっちを有効にするなら、下のst.write('エンターキーも押して消えます')も有効に
def kotae(a):
    if a == 1:
        st.success(f'正解!:{Ans}')
    elif a == 2:
        st.info(f'不正解!答えは{Ans}でした')
    st.write('エンターキーを押しても消えます')
    print(f'答え確認の時{st.session_state.count}')
    st.session_state.count += 1

    
#作成した問題を読み込んで回答させる。
with st.form(key="my_form",clear_on_submit=True):
    q,Ans = QandA(English,Japanese)
    st.subheader(f"{q} ({Average[Q]})")
    st.session_state.answer = Ans
    An = st.text_input("英語を入力してください",key="my_text")
    gtts.gTTS(Ans,lang='en').save("eitango.mp3")
    st.audio("eitango.mp3")
    submit_button=st.form_submit_button(label='答え合わせ',on_click=check)

wb.close()

